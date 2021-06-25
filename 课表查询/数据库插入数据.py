# -*- coding:UTF-8 -*-

import psycopg2
import openpyxl
# Connect to an existing database


conn = psycopg2.connect(host='localhost', 
	database="postgres", user="postgres", password="pass")

cur = conn.cursor()
cur.execute('''
DROP TABLE IF EXISTS course;
CREATE TABLE IF NOT EXISTS course  (
    sn       VARCHAR(20),     --序号
    no       VARCHAR(20), --课程号
    name     TEXT,        --课程名称
    teacher  TEXT,
    data     TEXT,        --课程时间
    place    TEXT,        --上课地点
    PRIMARY KEY(sn));
''')

cur.execute('''
DROP TABLE IF EXISTS student;
CREATE TABLE IF NOT EXISTS student (
    sn       VARCHAR(20),     --序号
    no       VARCHAR(20), --学号
    name     TEXT,        --姓名
    clss    TEXT,        --班级
    PRIMARY KEY(sn));
''')


cur.execute('''
DROP TABLE IF EXISTS course_grade;
CREATE TABLE IF NOT EXISTS course_grade (
    stu_sn VARCHAR(20),     -- 学生序号
    cou_sn VARCHAR(20),     -- 课程序号
    grade  NUMERIC(5,2), -- 最终成绩
    PRIMARY KEY(stu_sn, cou_sn));
''')


wb=openpyxl.load_workbook("数据.xlsx")
ws=wb.active

colC=ws["C"]     #name
colD=ws["D"]     #clss
colF=ws["F"]     #no
colG=ws["G"]     #课程name
colH=ws["H"]     #teacher
colI=ws["I"]     #data 学期
colK=ws["K"]     #stu_sn
colL=ws["L"]     #cou_sn
colM=ws["M"]     #grade
colN=ws["N"]     #place



for i in range(1,133):
    sn = i
    no=1910610000+i
    name = '%s' % colC[i].value
    clss = '%s' % colD[i].value
    place=i
    cur.execute('''
		INSERT INTO student(sn, no,name,clss) VALUES (%(no)s, %(sn)s,%(name)s,%(clss)s) 
	''', {'sn':sn, 'no':no,'name':name,"clss":clss} )

for i in range(1,31):
    sn = i
    no= '%s' % colF[i].value
    name = '%s' % colG[i].value
    teacher = '%s' % colH[i].value
    data = '%s' % colI[i].value
    cur.execute('''
		INSERT INTO course(sn, no,name,teacher,data,place) VALUES (%(sn)s, %(no)s,%(name)s,%(teacher)s,%(data)s,%(place)s) 
	''', {'sn':sn, 'no':no,'name':name,"teacher":teacher,"data":data,"place":place} )

for i in range(1,len(colK)):
    stu_sn = '%s' % colK[i].value
    cou_sn= '%s' % colL[i].value
    grade='%s' % colM[i].value
    cur.execute('''
		INSERT INTO course_grade(stu_sn,cou_sn,grade) VALUES (%(stu_sn)s,%(cou_sn)s,%(grade)s) 
	''', {'stu_sn':stu_sn, 'cou_sn':cou_sn,'grade':grade} )



conn.commit()